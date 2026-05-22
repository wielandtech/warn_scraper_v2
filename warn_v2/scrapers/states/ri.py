"""Rhode Island WARN scraper.

Source: https://dlt.ri.gov/employers/worker-adjustment-and-retraining-notification-warn
Data:   https://dlt.ri.gov/sites/g/files/xkgbur571/files/2025-06/Warn%20Report.xlsx

Schema (Excel, all-years workbook with sheets: 2026, 2025, 2024, Previous Years):
  WARN Date | Date Received | Company Name | Location of Layoffs |
  Number Affected | Effective Date | Closing Yes/No | Union Yes/No | Union Address

Each sheet starts with a title row, two blank rows, then the header row, then data.
"Closing Yes/No" = "Yes" maps to closure_type "Closure".
"Number Affected" can be an integer (native) or a messy string like
"9,891 Remote Workers (2 from RI)"; we extract the first integer found.
"""
from __future__ import annotations

import io
import re

import httpx
import openpyxl

from warn_v2.scrapers._helpers import as_date, as_str
from warn_v2.scrapers.base import NoticeRow, ParseFailed, ScrapeFailed
from warn_v2.scrapers.registry import register

PAGE_URL = (
    "https://dlt.ri.gov/employers/worker-adjustment-and-retraining-notification-warn"
)
SOURCE_URL = (
    "https://dlt.ri.gov/sites/g/files/xkgbur571/files/2025-06/Warn%20Report.xlsx"
)

_UA = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) warn-v2/0.1"
    )
}

_LEADING_INT = re.compile(r"\d[\d,]*")

_DATA_SHEETS = ("2026", "2025", "2024", "Previous Years")


class RIScraper:
    state = "RI"
    source_url = SOURCE_URL
    expected_row_range = (5, 5_000)
    required_fields = frozenset({"employer", "notice_date"})

    def fetch(self) -> bytes:
        try:
            r = httpx.get(SOURCE_URL, headers=_UA, timeout=60, follow_redirects=True)
            r.raise_for_status()
            return r.content
        except httpx.HTTPError as e:
            raise ScrapeFailed(f"GET {SOURCE_URL}: {e}") from e

    def parse(self, raw: bytes) -> list[NoticeRow]:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw))
        except Exception as e:
            raise ParseFailed(f"RI XLSX: could not open workbook: {e}") from e

        rows: list[NoticeRow] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows.extend(_parse_sheet(ws))

        if not rows:
            raise ParseFailed("RI XLSX: no data rows found in any sheet")
        return rows


def _parse_sheet(ws) -> list[NoticeRow]:
    all_rows = list(ws.iter_rows(values_only=True))
    # Find the header row: first row where first cell matches "warn date"
    header_idx: int | None = None
    for i, row in enumerate(all_rows):
        first = str(row[0]).strip().lower() if row[0] is not None else ""
        if "warn date" in first or first == "warn date":
            header_idx = i
            break
    if header_idx is None:
        return []

    header = [
        str(h).strip().lower().rstrip(" *").rstrip() if h is not None else ""
        for h in all_rows[header_idx]
    ]
    # Normalize trailing-space column names like "company name "
    header = [h.rstrip() for h in header]

    col: dict[str, int] = {name: i for i, name in enumerate(header) if name}

    # Find Company Name col (may have asterisk/note suffix in older sheets)
    company_col = next(
        (col[k] for k in col if k.startswith("company name")), col.get("company name")
    )
    if company_col is None:
        return []

    warn_date_col = col.get("warn date")
    if warn_date_col is None:
        return []

    rows: list[NoticeRow] = []
    for raw_row in all_rows[header_idx + 1 :]:
        employer = as_str(raw_row[company_col])
        if not employer:
            continue
        notice_date = as_date(raw_row[warn_date_col])
        if notice_date is None:
            continue

        eff_col = col.get("effective date")
        loc_col = col.get("location of layoffs")
        num_col = col.get("number affected")
        closing_col = col.get("closing yes/no")

        location_raw = as_str(raw_row[loc_col]) if loc_col is not None else None
        city = as_str(location_raw.split(",")[0].strip()) if location_raw else None

        count_raw = raw_row[num_col] if num_col is not None else None
        layoff_count = _parse_count(count_raw)

        closing_val = (
            str(raw_row[closing_col]).strip().lower() if closing_col is not None else ""
        )
        closure_type = "Closure" if closing_val == "yes" else None

        rows.append(
            NoticeRow(
                state="RI",
                employer=employer,
                notice_date=notice_date,
                effective_date=(
                    as_date(raw_row[eff_col]) if eff_col is not None else None
                ),
                layoff_count=layoff_count,
                closure_type=closure_type,
                city=city,
                source_url=SOURCE_URL,
            )
        )
    return rows


def _parse_count(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value) if not _is_nan(value) else None
    s = str(value)
    m = _LEADING_INT.search(s)
    if m:
        return int(m.group().replace(",", ""))
    return None


def _is_nan(v: float) -> bool:
    return v != v  # standard NaN check


register(RIScraper())
